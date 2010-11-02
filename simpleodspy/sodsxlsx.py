#!/usr/bin/env python
# -*- coding: utf-8 -*-

# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA 02111-1307, USA.

# Copyright (C) 2010 Yaacov Zamir (2010) <kzamir@walla.co.il>
# Author: Yaacov Zamir (2010) <kzamir@walla.co.il>

import re
from datetime import datetime

from openpyxl.workbook import Workbook
from openpyxl.style import Color, Border, Fill
from openpyxl.writer.excel import ExcelWriter

class SodsXlsx():
	def __init__(self, table, i_max = 30, j_max = 30):
		''' init and set default values for spreadsheet elements '''
		
		self.table = table
	
	def convertXlsBorderWidth(self, border):
		''' return the xls border width index '''
		
		# find the width in pt
		try:
			width = int(re.search('(.+)pt', border).group(1))
		except:
			return Border.BORDER_NONE
			
		# conver to excel widths
		if width > 2: xlsborder = Border.BORDER_THICK
		elif width == 2: xlsborder = Border.BORDER_MEDIUM
		elif width == 1: xlsborder = Border.BORDER_THIN
		else: xlsborder = Border.BORDER_NONE
		
		return xlsborder
	
	def convertXlsBorderColor(self, color_str):
		''' return the xls color index '''
		
		#find color #rgb in string
		try:
			color = re.search('#(......)', color_str).group(1)
		except:
			return Color('00000000')
		
		return Color('00' + color.upper()[1:])
		
	def saveXlsx(self, filename, i_max = None, j_max = None):
		''' save table in ods format '''
		
		if not i_max: i_max = self.table.i_max
		if not j_max: j_max = self.table.j_max
		
		# create new xlsx spreadsheet
		wb = Workbook()
		ew = ExcelWriter(workbook = wb)
		dest_filename = filename
		ws = wb.worksheets[0]
		ws.title = filename

		# make sure values are up to date
		# loop and update the cells value
		for i in range(1, i_max):
			for j in range(1, j_max):
				# update the cell text and condition
				cell = self.table.encodeColName(j) + str(i)
				self.table.updateOneCell(cell)
				c = self.table.getCellAt(i, j)
				
				# FIXME: excel output does not support conditional formating,
				# we do fixed formating of the conditional formating
				color = [c.color, c.condition_color][c.condition_state]
				background_color = [c.background_color, c.condition_background_color][c.condition_state]
				
				ws.cell(cell).style.font.name = c.font_family
				ws.cell(cell).style.font.size = int(c.font_size[:-2])
				ws.cell(cell).style.font.color = Color('FF' + color.upper()[1:])

				ws.cell(cell).style.fill.fill_type = 'solid'
				ws.cell(cell).style.fill.start_color = Color('00' + background_color.upper()[1:])
				ws.cell(cell).style.fill.end_color = Color('00' + background_color.upper()[1:])

				ws.cell(cell).style.borders.left.border_style = self.convertXlsBorderWidth(c.border_left)
				ws.cell(cell).style.borders.left.color = self.convertXlsBorderColor(c.border_left)
				ws.cell(cell).style.borders.right.border_style = self.convertXlsBorderWidth(c.border_right)
				ws.cell(cell).style.borders.left.color = self.convertXlsBorderColor(c.border_right)
				ws.cell(cell).style.borders.top.border_style = self.convertXlsBorderWidth(c.border_top)
				ws.cell(cell).style.borders.left.color = self.convertXlsBorderColor(c.border_top)
				ws.cell(cell).style.borders.bottom.border_style = self.convertXlsBorderWidth(c.border_bottom)
				ws.cell(cell).style.borders.left.color = self.convertXlsBorderColor(c.border_bottom)
				
				# set xls text
				if (c.formula):
					ws.cell(cell).value = c.formula
				elif c.value_type == 'float':
					ws.cell(cell).value = c.value
				elif c.value_type == 'date':
					ws.cell(cell).value = datetime.strptime(c.date_value, "%Y-%m-%d")
				else:
					ws.cell(cell).value = c.text + " "
		
		ew.save(filename = filename)
		
if __name__ == "__main__":
	
	from sodsspreadsheet import SodsSpreadSheet
	
	t = SodsSpreadSheet()
	
	print "Test spreadsheet naming:"
	print "-----------------------"
	
	t.setStyle("A1", text = "Simple ods python")
	t.setStyle("A1", font_size = "33pt")
	t.setStyle("D2", font_size = "23pt", color = "#ff00ff")
	t.setStyle("A1", background_color = "#00ff00")
	t.setStyle("A2", background_color = "#ffff00")
	t.setStyle("A3", background_color = "#0000ff")
	
	t.setValue("A2", 123.4)
	t.setValue("B2", "2010-01-01")
	t.setValue("C2", "=0.6")
	t.setValue("D2", "= A2 + 3")
	
	t.setStyle("A3:D3", border_top = "1pt solid #ff0000")
	t.setValue("C3", "Sum of cells:")
	t.setValue("D3", "=sum(A2:D2)")
	
	t.setStyle("D2:D3", condition = "cell-content()<=200")
	t.setStyle("D2:D3", condition_color = "#ff0000")
	
	tw = SodsXlsx(t)
	tw.saveXlsx("test.xlsx")
	
